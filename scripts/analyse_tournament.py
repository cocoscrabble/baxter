#!/usr/bin/env python
"""Load a tournament results spreadsheet into a PairingData for the pairing engine.

The spreadsheet is the community pairing workbook; we read five of its tabs
(Entrants, Results, Settings, FixedPairing, Pairings) and build the same
PairingData the Django app builds from the ORM, so the pairing code can be run
against real tournament data without touching the database.

Usage: uv run python scripts/analyse_tournament.py ~/tmp/results/tourney.xlsx
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl

# Ensure the project root is on sys.path when run from any directory.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from tournaments.pairing.base import (  # noqa: E402
    EntrantData,
    PairingData,
    PlayerData,
    Repeats,
    ResultSlipData,
    standings_after_round,
)
from tournaments.pairing.round_pairing import RP, RoundPairing  # noqa: E402

# The only key/value settings the sheet defines. Anything else in that range is
# not a setting — older workbooks put the strategy legend there.
SWISS_SETTING_KEYS = {"swiss_weight", "swiss_distance"}

# The Results tab is a form-response dump: nine meaningful columns followed by a
# wide band of empty ones, and a max_row that runs far past the last entry.
RESULT_COLUMNS = 9

# The workbook seeds itself with a placeholder entrant whose round-1 bye keeps
# the response form alive ("PLEASE DO NOT DELETE THE TEST PLAYER RESULT"). It is
# not a competitor, so it is dropped along with its bye on import.
EXCLUDED_PLAYERS = {"Test Player"}

# The workbook's strategy codes (documented in column G of the Settings tab) are
# a different vocabulary from Baxter's ABBREV: they name Swiss variants by which
# round's standings they pair off ("S" = last round, "ST" = two rounds ago),
# where Baxter carries that in RoundPairing.start_round. Each entry maps a sheet
# code to (strategy, pair_from), pair_from being how many rounds back the
# standings come from.
STRATEGY_CODES: dict[str, tuple[RP, int]] = {
    "QD": (RP.Quads_Distributed, 1),
    "QC": (RP.Quads_Clustered, 1),
    "QE": (RP.Quads_Equalized, 1),
    "SIX": (RP.Sixes, 1),
    "K": (RP.KotH, 1),
    "Q": (RP.QotH, 1),
    "R": (RP.RoundRobin, 1),
    "DR": (RP.DoubleRoundRobin, 1),
    "CH": (RP.Charlottesville, 1),
    "S": (RP.Swiss, 1),
    "ST": (RP.Swiss, 2),
    "RAND": (RP.Random, 1),
    "RANDNR": (RP.RandomNoRepeats, 1),
    "SPR": (RP.SwissPlusRandom, 1),
    "STPR": (RP.SwissPlusRandom, 2),
}

# Quad/hex blocks are numbered per set in the sheet (QD1, QD1, QD1; QD2, ...);
# the trailing digits identify the set, not a distinct strategy.
_CODE_SUFFIX = re.compile(r"^([A-Z]+?)(\d*)$")

# A FixedPairing cell is either a player name or "#n" (the nth place in the
# standings the round pairs off). Names in the wild carry a readability
# annotation — "Josh Sokol (#10)" — which is not part of the name.
_STANDINGS_REF = re.compile(r"^#(\d+)$")
_NAME_ANNOTATION = re.compile(r"\s*\([^)]*\)\s*$")


@dataclass
class Entrant:
    """One row of the Entrants tab."""

    name: str
    rating: int
    seed: int | None = None


@dataclass
class Result:
    """One reported game from the Results tab."""

    submitted_on: datetime | None
    round: int
    winner: str
    winner_score: int
    opponent: str
    opponent_score: int
    winner_went: str | None  # "First" / "Second"
    affirmation: str | None

    @property
    def is_bye(self) -> bool:
        return "bye" in (self.winner.lower(), self.opponent.lower())

    @property
    def is_tie(self) -> bool:
        return self.winner_score == self.opponent_score

    @property
    def winner_started(self) -> bool:
        return (self.winner_went or "").lower() == "first"

    @property
    def spread(self) -> int:
        return self.winner_score - self.opponent_score


@dataclass
class FixedPair:
    """One row of the FixedPairing tab.

    player1/player2 are kept as written: either a name or a "#n" standings
    reference, resolved to names later against the standings the round pairs off.
    """

    round: int
    player1: str
    player2: str
    force_player1_start: bool = False


@dataclass
class Settings:
    """The Settings tab: a round->strategy schedule plus a key/value block."""

    # round -> raw sheet code, e.g. {1: "QE", 4: "STPR"}
    round_codes: dict[int, str] = field(default_factory=dict)
    swiss: dict[str, object] = field(default_factory=dict)

    @property
    def rounds(self) -> int:
        return max(self.round_codes, default=0)


@dataclass
class Tournament:
    path: Path
    entrants: list[Entrant]
    results: list[Result]
    settings: Settings
    fixed_pairs: list[FixedPair] = field(default_factory=list)
    # Round -> the pairings the sheet's own script generated (Pairings tab).
    generated_pairings: dict[int, list[tuple[str, str]]] = field(default_factory=dict)

    @property
    def players(self) -> set[str]:
        names = {r.winner for r in self.results} | {r.opponent for r in self.results}
        return {n for n in names if n.lower() != "bye"}

    def results_by_round(self) -> dict[int, list[Result]]:
        rounds: dict[int, list[Result]] = {}
        for r in self.results:
            rounds.setdefault(r.round, []).append(r)
        return dict(sorted(rounds.items()))


def _int(value) -> int | None:
    """Coerce a cell to int. Numeric cells arrive as floats (1.0, 457.0)."""
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _str(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def load_entrants(sheet) -> list[Entrant]:
    entrants = []
    for row in sheet.iter_rows(min_row=2, max_col=5, values_only=True):
        name = _str(row[0])
        if not name or name in EXCLUDED_PLAYERS:
            continue
        entrants.append(Entrant(name=name, rating=_int(row[1]) or 0, seed=_int(row[4])))
    return entrants


def load_results(sheet) -> list[Result]:
    rows = sheet.iter_rows(min_row=2, max_col=RESULT_COLUMNS, values_only=True)
    results = []
    for row in rows:
        submitted, rnd, winner, wscore, opponent, oscore, went, affirm = row[:8]
        # max_row overshoots the data; skip the empty tail and any interior gaps.
        if not any(v is not None for v in row[:8]):
            continue
        round_no = _int(rnd)
        if round_no is None:
            continue
        results.append(
            Result(
                submitted_on=submitted if isinstance(submitted, datetime) else None,
                round=round_no,
                winner=_str(winner) or "",
                winner_score=_int(wscore) or 0,
                opponent=_str(opponent) or "",
                opponent_score=_int(oscore) or 0,
                winner_went=_str(went),
                affirmation=_str(affirm),
            )
        )
    return [
        r
        for r in results
        if not (EXCLUDED_PLAYERS & {r.winner, r.opponent}) and not r.is_bye
    ]


def load_fixed_pairs(sheet) -> list[FixedPair]:
    pairs = []
    # Columns A-D; column F onwards documents the format.
    for rnd, p1, p2, force in sheet.iter_rows(min_row=2, max_col=4, values_only=True):
        round_no = _int(rnd)
        name1, name2 = _str(p1), _str(p2)
        if round_no is None or not name1 or not name2:
            continue
        pairs.append(
            FixedPair(
                round=round_no,
                player1=name1,
                player2=name2,
                force_player1_start=bool(_str(force)),
            )
        )
    return pairs


def load_generated_pairings(sheet) -> dict[int, list[tuple[str, str]]]:
    """The pairings the sheet's own script produced, from the Pairings tab.

    Laid out as repeating "ROUND n" / "Table | 1st Player | 2nd Player" blocks.
    Names carry the "(#10)" seed annotation. Order is table order, not standings
    order, so only the pairs themselves are meaningful.
    """
    rounds: dict[int, list[tuple[str, str]]] = {}
    current = None
    for label, p1, p2 in sheet.iter_rows(min_row=1, max_col=3, values_only=True):
        header = _str(label)
        match = re.match(r"^ROUND\s+(\d+)", header) if header else None
        if match:
            current = int(match.group(1))
            rounds.setdefault(current, [])
            continue
        if current is None or header == "Table":
            continue
        name1, name2 = _str(p1), _str(p2)
        if not name1 or not name2:
            continue
        rounds[current].append(
            (_NAME_ANNOTATION.sub("", name1).strip(), _NAME_ANNOTATION.sub("", name2).strip())
        )
    return {r: p for r, p in rounds.items() if p}


def load_settings(sheet) -> Settings:
    settings = Settings()
    # Columns A/B hold the per-round pairing schedule and D/E the swiss key/value
    # block; everything from G rightwards is human documentation.
    rows = sheet.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True)
    for rnd, pairing in rows:
        round_no = _int(rnd)
        code = _str(pairing)
        if round_no is not None and code:
            settings.round_codes[round_no] = code.upper()

    kv = sheet.iter_rows(min_row=2, min_col=4, max_col=5, values_only=True)
    for key, value in kv:
        name = _str(key)
        # Only the known keys: on an older workbook this range holds the strategy
        # legend rather than a key/value block, and every line of prose would
        # otherwise be read in as a setting.
        if name not in SWISS_SETTING_KEYS:
            continue
        settings.swiss[name] = _int(value) if _int(value) is not None else _str(value)

    return settings


def load_tournament(path: Path) -> Tournament:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        required = {"Entrants", "Results", "FixedPairing", "Pairings"}
        missing = required - set(workbook.sheetnames)
        if missing:
            names = ", ".join(sorted(missing))
            raise ValueError(f"{path.name} is missing tab(s): {names}")
        # Older workbooks keep the round schedule on a "RoundPairing" tab and have
        # no swiss settings at all (that version of the script hardcoded them).
        # The tab layout is otherwise identical, so either name works.
        schedule = next(
            (n for n in ("Settings", "RoundPairing") if n in workbook.sheetnames), None
        )
        if schedule is None:
            raise ValueError(
                f"{path.name} has neither a Settings nor a RoundPairing tab"
            )
        return Tournament(
            path=path,
            entrants=load_entrants(workbook["Entrants"]),
            results=load_results(workbook["Results"]),
            settings=load_settings(workbook[schedule]),
            fixed_pairs=load_fixed_pairs(workbook["FixedPairing"]),
            generated_pairings=load_generated_pairings(workbook["Pairings"]),
        )
    finally:
        workbook.close()


# ---------------------------------------------------------------------------
# Conversion to the pairing engine's input
# ---------------------------------------------------------------------------


def parse_strategy_code(code: str) -> tuple[RP, int]:
    """Map a sheet strategy code to (strategy, pair_from), ignoring a set number."""
    match = _CODE_SUFFIX.match(code.upper())
    base = match.group(1) if match else code.upper()
    if base not in STRATEGY_CODES:
        raise ValueError(f"unknown pairing code {code!r}")
    return STRATEGY_CODES[base]


def to_round_pairings(settings: Settings) -> list[RoundPairing]:
    """Build the per-round schedule, resolving each round's standings source.

    start_round follows the same family rules as blocks_to_round_pairings: a
    round-robin block rotates off a fixed order (start_round = block start), a
    quad/hex block pairs off one snapshot taken before the block, and sliding
    strategies pair off pair_from rounds back.
    """
    rounds = sorted(settings.round_codes)
    out: list[RoundPairing] = []
    prev_code = None
    start = 1
    for round_no in rounds:
        code = settings.round_codes[round_no]
        strategy, pair_from = parse_strategy_code(code)
        # A change of code starts a new block; the set number in QD1/QD2 is what
        # separates two adjacent blocks of the same strategy.
        if code != prev_code:
            start = round_no
        prev_code = code

        if RP.is_round_robin(strategy):
            start_round = start
        elif RP.is_quad(strategy):
            start_round = start - pair_from
        else:
            start_round = round_no - pair_from
        out.append(RoundPairing(round_no, max(start_round, 0), str(strategy)))
    return out


def resolve_fixed_pairings(
    pd: PairingData, fixed_pairs: list[FixedPair], known: set[str]
) -> dict[int, list[tuple[str, str]]]:
    """Resolve FixedPairing rows to the name pairs PairingData carries.

    A "#n" cell names the nth player in the standings the round pairs off, so it
    is resolved against standings_after_round at that round's start_round — the
    same snapshot the strategy will see.
    """
    schedule = {rp.round: rp for rp in pd.round_pairings}
    standings_cache: dict[int, list[str]] = {}

    def resolve(cell: str, round_no: int) -> str:
        ref = _STANDINGS_REF.match(cell)
        if not ref:
            # Strip the "(#10)" readability annotation only if it isn't the name.
            if cell in known:
                return cell
            stripped = _NAME_ANNOTATION.sub("", cell).strip()
            if stripped not in known:
                raise ValueError(f"round {round_no}: unknown player {cell!r}")
            return stripped
        rp = schedule.get(round_no)
        start_round = rp.start_round if rp else round_no - 1
        if start_round not in standings_cache:
            standings_cache[start_round] = [
                p.name for p in standings_after_round(pd, start_round)
            ]
        standings = standings_cache[start_round]
        place = int(ref.group(1))
        if not 1 <= place <= len(standings):
            raise ValueError(
                f"round {round_no}: {cell} is out of range for a "
                f"{len(standings)}-player field"
            )
        return standings[place - 1]

    fixed: dict[int, list[tuple[str, str]]] = {}
    for fp in fixed_pairs:
        pair = (resolve(fp.player1, fp.round), resolve(fp.player2, fp.round))
        fixed.setdefault(fp.round, []).append(pair)
    return fixed


def fixed_starts(
    fixed: dict[int, list[tuple[str, str]]], fixed_pairs: list[FixedPair]
) -> dict[tuple[int, str], bool]:
    """(round, name) -> True for rows with "Force Player1 Start" set.

    PairingData carries no forced-start field and the engine's JSON boundary has
    none either, so this is returned alongside for use with Starts(fixed_starts=).
    """
    starts = {}
    for fp in fixed_pairs:
        if not fp.force_player1_start:
            continue
        for pair in fixed.get(fp.round, []):
            first = pair[0]
            if first in (fp.player1, _NAME_ANNOTATION.sub("", fp.player1).strip()):
                starts[(fp.round, first)] = True
    return starts


def to_swiss_config(settings: Settings) -> dict:
    """Map the sheet's Swiss settings onto Baxter's swiss_config.

    The sheet spends swiss_distance twice: pairCandidates drops candidate edges
    whose distance is >= it, and pairSwissPlusRandom slices the Swiss-paired top
    that many players deep. Baxter keeps those as separate knobs (max_distance,
    spr_split), so both are set from the one sheet value to match its behaviour.
    """
    weight = settings.swiss.get("swiss_weight")
    distance = settings.swiss.get("swiss_distance")
    config: dict = {}
    if isinstance(weight, int):
        config["swiss_weight"] = weight
    if isinstance(distance, int):
        config["max_distance"] = distance
        config["spr_split"] = distance
    return config


def to_pairing_data(tournament: Tournament) -> PairingData:
    """Build the PairingData the pairing engine consumes."""
    entrants = [
        EntrantData(PlayerData(name=e.name, rating=e.rating))
        for e in tournament.entrants
    ]
    # Entrants who only ever appear in the results (not on the Entrants tab) would
    # otherwise be invisible to seedings; add them unrated so they still pair.
    known = {e.player.name for e in entrants}
    for name in sorted(tournament.players - known):
        entrants.append(EntrantData(PlayerData(name=name, rating=0)))

    slips = [
        ResultSlipData(
            round=r.round,
            winner_name=r.winner,
            loser_name=r.opponent,
            winner_score=r.winner_score,
            loser_score=r.opponent_score,
            winner_started=r.winner_started,
        )
        for r in tournament.results
    ]
    pd = PairingData(
        result_slips=slips,
        entrants=entrants,
        repeats=Repeats(),
        round_pairings=to_round_pairings(tournament.settings),
        swiss_config=to_swiss_config(tournament.settings) or None,
    )
    known = {e.player.name for e in pd.entrants}
    pd.fixed_pairings = resolve_fixed_pairings(pd, tournament.fixed_pairs, known)
    return pd


def summarise(tournament: Tournament) -> None:
    settings = tournament.settings
    by_round = tournament.results_by_round()
    pd = to_pairing_data(tournament)
    schedule = {rp.round: rp for rp in pd.round_pairings}

    print(f"{tournament.path.name}")
    print(f"  entrants: {len(tournament.entrants)}")
    print(f"  results:  {len(tournament.results)} over {len(by_round)} rounds")
    print(f"  rounds scheduled: {settings.rounds}")
    if settings.swiss:
        pairs = ", ".join(f"{k}={v}" for k, v in settings.swiss.items())
        print(f"  swiss settings:   {pairs}")

    unplayed = tournament.players - {e.name for e in tournament.entrants}
    if unplayed:
        print(f"  !! in results but not on Entrants tab: {sorted(unplayed)}")

    print("\n  round  code    strategy             from  games  ties  fixed")
    for round_no in range(1, max(settings.rounds, max(by_round, default=0)) + 1):
        games = by_round.get(round_no, [])
        code = settings.round_codes.get(round_no, "-")
        rp = schedule.get(round_no)
        strategy = rp.pairing if rp else "-"
        source = str(rp.start_round) if rp else "-"
        ties = sum(g.is_tie for g in games)
        fixed = len(pd.fixed_pairings.get(round_no, []))
        print(
            f"  {round_no:5d}  {code:<6}  {strategy:<20} {source:>4}  "
            f"{len(games):5d}  {ties:4d}  {fixed or '':>5}"
        )

    if pd.fixed_pairings:
        print("\n  fixed pairings:")
        for round_no, pairs in sorted(pd.fixed_pairings.items()):
            for first, second in pairs:
                print(f"    R{round_no}: {first} vs {second}")
    starts = fixed_starts(pd.fixed_pairings, tournament.fixed_pairs)
    if starts:
        forced = ", ".join(f"R{r} {n}" for (r, n) in sorted(starts))
        print(f"  forced starts (not carried by PairingData): {forced}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("path", type=Path, help="path to the .xlsx results file")
    args = parser.parse_args()

    tournament = load_tournament(args.path.expanduser())
    summarise(tournament)


if __name__ == "__main__":
    main()
