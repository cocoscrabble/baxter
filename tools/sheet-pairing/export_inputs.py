#!/usr/bin/env python
"""Dump a tournament workbook's inputs as a JavaScript literal module.

The Apps Script reads five ranges off the spreadsheet and hands the raw rows to
its `make*` builders. This writes exactly those rows out as JS literals, so the
standalone runner can feed the real parsers the same arrays the sheet would have,
with no spreadsheet in the loop.

Ranges mirror the `collect*` functions in Code.ts:

    Results      B2:H   round, winner, winner score, loser, loser score, first?
    Entrants     A2:E   name, rating, _, table, seed
    Settings     A2:B   round, pairing code
    Settings     D2:E   setting name, value
    FixedPairing A2:D   round, player1, player2, force player1 start

Usage: uv run python tools/sheet-pairing/export_inputs.py tourney.xlsx -o out.js
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import sys
from pathlib import Path

import openpyxl

# (attribute name, sheet, first column, last column) — 1-based, inclusive.
RANGES = [
    ("results", "Results", 2, 8),
    ("entrants", "Entrants", 1, 5),
    ("roundPairings", "Settings", 1, 2),
    ("settings", "Settings", 4, 5),
    ("fixedPairings", "FixedPairing", 1, 4),
]

# Older workbooks keep the schedule on a "RoundPairing" tab and carry no swiss
# settings block; the layout is otherwise the same.
SCHEDULE_TABS = ("Settings", "RoundPairing")


def cell(value):
    """Coerce a cell to what Apps Script's getValues() would hand the builders.

    Sheets hands back numbers, strings, or "" for a blank. openpyxl gives None
    for blanks and floats for whole numbers, so normalise both — the builders do
    `parseInt` and `.toLowerCase()` on these and would throw on None.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def read_range(workbook, sheet_name: str, first_col: int, last_col: int) -> list[list]:
    """Rows 2..n of a column range, trimmed at the last row with any content.

    `collectSheetData` stops at the end of the contiguous data region; the sheets
    declare far more rows than they use, so trim to the last non-empty row rather
    than emitting thousands of blanks.
    """
    sheet = workbook[sheet_name]
    rows = []
    for row in sheet.iter_rows(
        min_row=2, min_col=first_col, max_col=last_col, values_only=True
    ):
        rows.append([cell(v) for v in row])
    while rows and all(v == "" for v in rows[-1]):
        rows.pop()
    return rows


def restore_withdrawals(data: dict[str, list[list]]) -> list[str]:
    """Add back entrants the director deleted mid-tournament.

    The Entrants tab is a live control input, not a record: the sheet's only way
    to stop pairing a player who withdraws is to delete their row, so the tab
    ends up describing the field as it was at the *end*. Anyone who appears in
    Results but not on Entrants played and then withdrew, and without their row
    the sheet's own standings (getCurrentEntrantsRanking filters to entrants)
    silently omit them — so it cannot re-derive the rounds they played in.

    Ratings are left at 0 deliberately: they only affect round-1 seeding, and
    every round that pairs off standings is unaffected. `--active-at` in the
    runner is what puts each round's field back to what it was.

    Returns the names restored.
    """
    listed = {row[0] for row in data["entrants"] if row[0]}
    played: dict[str, int] = {}
    for row in data["results"]:
        try:
            rnd = int(row[0])
        except (TypeError, ValueError):
            continue
        for name in (row[1], row[3]):
            if name and name != "Bye":
                played[name] = max(played.get(name, 0), rnd)
    missing = sorted(n for n in played if n not in listed)
    seed = max((int(r[4]) for r in data["entrants"] if str(r[4]).strip().isdigit()), default=0)
    for name in missing:
        seed += 1
        data["entrants"].append([name, 0, "", "", seed])
    return missing


def export(path: Path) -> dict[str, list[list]]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        schedule = next((t for t in SCHEDULE_TABS if t in workbook.sheetnames), None)
        if schedule is None:
            raise SystemExit(f"{path.name} has none of {SCHEDULE_TABS}")
        needed = {s for _, s, _, _ in RANGES if s not in SCHEDULE_TABS}
        missing = needed - set(workbook.sheetnames)
        if missing:
            raise SystemExit(f"{path.name} is missing tab(s): {sorted(missing)}")
        return {
            name: read_range(workbook, schedule if sheet in SCHEDULE_TABS else sheet, lo, hi)
            for name, sheet, lo, hi in RANGES
        }
    finally:
        workbook.close()


def render(data: dict[str, list[list]], source: str) -> str:
    out = [
        "// GENERATED by tools/sheet-pairing/export_inputs.py — do not edit.",
        f"// Source workbook: {source}",
        "//",
        "// The five ranges the Apps Script reads, exactly as its `make*` builders",
        "// receive them. Edit a value here to explore a what-if.",
        "",
    ]
    for name, _, _, _ in RANGES:
        rows = data[name]
        out.append(f"// {len(rows)} rows")
        out.append(f"const {name} = [")
        for row in rows:
            out.append("  " + json.dumps(row, ensure_ascii=False) + ",")
        out.append("];")
        out.append("")
    names = ", ".join(name for name, _, _, _ in RANGES)
    out.append(f"module.exports = {{ {names}, source: {json.dumps(source)} }};")
    out.append("")
    return "\n".join(out)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("path", type=Path, help="path to the .xlsx workbook")
    parser.add_argument("-o", "--out", type=Path, help="output .js file (default stdout)")
    args = parser.parse_args()

    path = args.path.expanduser()
    data = export(path)
    restored = restore_withdrawals(data)
    if restored:
        print(f"restored {len(restored)} withdrawn entrant(s): {', '.join(restored)}",
              file=sys.stderr)
    text = render(data, path.name)
    if args.out:
        args.out.write_text(text)
        print(f"wrote {args.out}")
    else:
        print(text)


if __name__ == "__main__":
    main()
